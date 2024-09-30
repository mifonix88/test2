# -*- coding: utf-8 -*-
import csv
import sys
import os
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


class OpenConfig():
    def __init__(self, path_config):
        self.config = {}
        self.openCofig(path_config)

        self.path_to_exel = self.config['list0'].pop('Путь до EXCEL') 
        self.delimiter = self.config['list0'].pop('Разделитель')
        self.art = self.config['list0'].pop('Арт.')
        self.name = self.config['list0'].pop('Наименование')
        self.kol = self.config['list0'].pop('Всего:')
        self.id_ = self.config['list0'].pop('Отправления')
        self.config.pop('list0')

    def openCofig(self, path_config):
        num = 0
        resalt = {}
        name = None

        with open(path_config, encoding='utf-8') as _file:
            lins = [i for i in _file.readlines() if i != '\n']# список строк в котором есть сомволы отлисные от окончания строки

        for i in lins:
            if '#' not in i[0]: #если не коментарий
                f = i.find('=')                                     # Слещение знак = в строке 
                if f > 0:                                           # если знак = есть
                    name = i[:f].strip()                               # это параметр, получаем его имя
                    arg = self.listArg(i[f+1:])                         #получаем его параметры
                    if name != 'name':                              # если его имя не равно name #Записываем в словарь имя и параметры
                        if len(arg) == 1:                           #если параметр 1 записываем напрямую
                            resalt[name] =  arg[0]
                        else:
                            resalt[name] = arg                      #иначе пишим список
                        self.config[f'list{num}'] = resalt    #записываем словарь в словарь
                    else:
                        num+=1
                        resalt = {name:arg[0]}
                        self.config[f'list{num}'] = resalt   #если параметр name создаём новый словарь
                else:                                              #если знака = нет
                    arg = self.listArg(i)
                    if arg[0]:                                    #а список параметьров не пуст
                        self.config[f'list{num}'][name]+=arg #записываем его как доп параметры к последнему с именем

            
    def digitArg(self, arg):
        '''если str состоит только из чисел вернёт int'''
        return int(arg) if arg.isdigit() else  arg

    def listArg(self, arg):
        '''превращает str в list по разделителю "," отсекает симфол конца строки "\n" 
           Уберает не буквенные симвомы с лево и c права, целые числа превращает в int'''
        return [self.digitArg(j.strip()) for j in arg.split(',') if j != '\n']




class Reader(OpenConfig):
    def __init__(self,arg):
        super().__init__(arg)
        self.values = {} #art:name,[kol]
        self.id_values = {} #id:{art:[art], kol:[kol]}

    def readFile(self, file_name):
        with open(file_name, encoding='utf-8') as r_file:
            reader = csv.DictReader(r_file, delimiter = self.delimiter)
            
            for i in reader:

                art = self.digitArg(i[self.art])
                name = self.digitArg(i[self.name])
                kol = self.digitArg(i[self.kol])
                id_ = self.digitArg(i[self.id_])

                if self.values.get(art):
                    self.values[art][name].append(kol)
                else:
                    self.values[art] = {name:[kol]}

                #для сборныx
                if self.id_values.get(id_):  #если записываем
                    self.id_values[id_][art] = kol
                else:
                    self.id_values[id_] = {art:kol}


    def collectingOrders(self):
        '''вернёт список ключей из id_values которые подходят критерию
           сборного заказа'''
        res = []
        for i in self.id_values.keys():
            if len(self.id_values[i].keys()) > 1:
                res.append(i)
            else:
                if sum(self.id_values[i].values()) > 1:
                    res.append(i)
        return sorted(res)

class MyWorkbook(Reader):
    def __init__(self,arg):
        super().__init__(arg)


    def _top(self, ls, ws):
        for i in ls:
            arg = ws.cell(row=1, column=1+ls.index(i), value = i)
            arg.font = Font(size= 12, bold=True)
            arg.alignment = Alignment(horizontal='center')



    def addTitul(self):
        self.wb =  Workbook()
        ws = self.wb.active
        ws.title = "Общий"
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['A'].width = 20
        ls = ('Арт.', 'Наименование', 'Всего:', 'По 1шт', 'Сборные')
        self._top(ls, ws)
        

        key = sorted(self.values.keys())  #список артиклов

        collecting_or = self.collectingOrders() #список сборных отправлений

        collecting_art = {} #словарь {артикл:[количество]} из сборных отправлений
        
        temp_in = len(key) + 3

        for i in collecting_or:
            for j in self.id_values[i].keys():
                if collecting_art.get(j):
                    collecting_art[j].append(self.id_values[i][j])
                else:
                    collecting_art[j] = [self.id_values[i][j]]

        for index, art in enumerate(key):
            
            index = index + 2
            
            art_name = list(self.values[art].keys())[0]
            sum_val = sum(self.values[art][art_name])

            arg1 = ws.cell(row=index, column=1, value = art)
            arg1.alignment = Alignment(horizontal='center')
            
            arg2 = ws.cell(row=index, column=2, value = art_name)
            arg2.alignment = Alignment(horizontal='left')

            arg3 = ws.cell(row=index, column=3, value = sum_val)
            arg3.alignment = Alignment(horizontal='center')
            arg3.font = Font(bold=True)

            sb = []
            if art in collecting_art:
                sb = collecting_art[art]

            sb2 = str(sb) if sb else ''

            arg5 = ws.cell(row=index, column=5, value = sb2)
            arg5.alignment = Alignment(horizontal='center')
            
            su = sum_val - sum(sb)
            su = su if su else ''

            arg4 = ws.cell(row=index, column=4, value = su)
            arg4.alignment = Alignment(horizontal='center')


        arg = ws.cell(row=temp_in, column=2, value = f"Список сборных заказов (Всего: {len(collecting_or)})")
        arg.font = Font(size= 14, bold=True)
        arg.alignment = Alignment(horizontal='center')

        temp_in +=1
        
        ls = ('Отправления', 'Наименование', 'Кол-во:')
        for i in ls:
            arg = ws.cell(row=temp_in, column=1+ls.index(i), value = i)
            arg.font = Font(size= 12, bold=True)
            arg.alignment = Alignment(horizontal='center')


        for i in collecting_or:
            if len(self.id_values[i].keys()) > 1:
                temp_in +=2
            else:
                temp_in +=1

            a = ws.cell(row=temp_in, column=1, value = i)
            a.alignment = Alignment(horizontal='center')

            for j in self.id_values[i]:
                
                art , kol = j, self.id_values[i][j]
                name = list(self.values[art].keys())[0]

                b = ws.cell(row=temp_in, column=2, value = name)
                b.alignment = Alignment(horizontal='left')
                
                c = ws.cell(row=temp_in, column=3, value = kol)
                c.alignment = Alignment(horizontal='center')
                
                if len(self.id_values[i]) > 1:
                    temp_in +=1


    def addList(self, list_id):
        
        switch = False
        
        name = self.config[list_id]['name']
        artikl = self.config[list_id]['art']
        artikl = sorted(artikl) if type(artikl) != int else [artikl]
        
        for i in artikl:
            if self.values.get(i):
                switch = True
        
        if switch:
            ws = self.wb.create_sheet(name)
            ws.column_dimensions['B'].width = 80
            ws.column_dimensions['A'].width = 10

            ls = ('Арт.', 'Наименование', 'Всего:')
            self._top(ls, ws)
            
            index = 2
            for art in artikl:
                if self.values.get(art):
                    
                    arg1 = ws.cell(row=index, column=1, value = art)
                    arg1.alignment = Alignment(horizontal='center')
                    
                    art_name = list(self.values[art].keys())[0]

                    arg2 = ws.cell(row=index, column=2, value = art_name)
                    arg2.alignment = Alignment(horizontal='left')

                    arg3 = ws.cell(row=index, column=3, value = sum(self.values[art][art_name]))
                    arg3.alignment = Alignment(horizontal='center')
                    arg3.font = Font(bold=True)
                    index += 1

    def addCollecting(self):
        ws = self.wb.create_sheet('Список отправлений')
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['A'].width = 20

        index = 1
        sort_id = sorted(self.id_values)

        ls = (f'Отправления: {len(sort_id)}', 'Наименование', 'Кол-во:')
        self._top(ls, ws)
            
        for i in sort_id:
            if len(self.id_values[i].keys()) > 1:
                index +=2
            else:
                index +=1

            a = ws.cell(row=index, column=1, value = i)
            a.alignment = Alignment(horizontal='center')

            for j in self.id_values[i]:
                
                art , kol = j, self.id_values[i][j]
                name = list(self.values[art].keys())[0]

                b = ws.cell(row=index, column=2, value = name)
                b.alignment = Alignment(horizontal='left')
                
                c = ws.cell(row=index, column=3, value = kol)
                c.alignment = Alignment(horizontal='center')
                
                if len(self.id_values[i]) > 1:
                    index +=1


    def saveFile(self, name):

        self.addTitul()
        for i in self.config.keys():
            self.addList(i)
        self.addCollecting()

        try:
            os.remove(f'{name}.xlsx')
        except(FileNotFoundError):pass

        self.wb.save(f'{name}.xlsx')


def path_to_root(name):
    '''принемает путь до открываемого файла (str)
       вернёт путь для сохранения одноимённого файла в корень программы(str)'''
    name_file = name[:-4].split('\\')[-1]#имя файла без расширения(-4)
    root = sys.argv[0].split('\\')[:-1]
    root.append(name_file)
    name = '\\'.join(root)
    return name

def open_exel(path, name):
    '''Открывает файл excel по указаному адресу'''
    subprocess.Popen([path, f'{name}.xlsx'])



def main():
    try:
        path = sys.argv[1]#получаем путь к открываемому файлу

        config = path_to_root('config....')
        book = MyWorkbook(config)#передаём кофиг
    
        book.readFile(path)#Получаем данные из файла

        path = path_to_root(path)
        book.saveFile(path)#записываем полученые данные в файл EXEL

        open_exel(book.path_to_exel, path)
    except(IndexError):pass


if __name__ == '__main__':
    main()
